import openpyxl
from openpyxl import Workbook


try:
    wb = openpyxl.load_workbook("informe_gastos.xlsx")
except FileNotFoundError:

    wb = Workbook("informe_gastos.xlsx")
    wb.save("informe_gastos.xlsx")


wb = openpyxl.load_workbook("informe_gastos.xlsx")


if "Gastos" in wb.sheetnames:
    worksheet = wb["Gastos"]
else:
    worksheet = wb.active
    worksheet.title = "Gastos"


def ingresar_gasto():
    fecha = input("Fecha del gasto (YYYY-MM-DD): ")
    descripcion = input("Descripción del gasto: ")
    monto = float(input("Monto del gasto: $"))
    return fecha, descripcion, monto


while True:
    fecha, descripcion, monto = ingresar_gasto()
    worksheet.append([fecha, descripcion, monto])
    continuar = input("¿Deseas ingresar otro gasto? (S/N): ").strip().lower()
    if continuar != 's':
        break

num_gastos = worksheet.max_row - 1 
gastos = list(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=3))
monto_total = sum([gasto[2].value for gasto in gastos])
gasto_maximo = max(gastos, key=lambda x: x[2].value)
gasto_minimo = min(gastos, key=lambda x: x[2].value)


print("\nResumen de gastos:")
print(f"Número total de gastos: {num_gastos}")
print(f"Gasto más caro: Fecha: {gasto_maximo[0].value} - Descripción: {gasto_maximo[1].value} - Monto: ${gasto_maximo[2].value}")
print(f"Gasto más barato: Fecha: {gasto_minimo[0].value} - Descripción: {gasto_minimo[1].value} - Monto: ${gasto_minimo[2].value}")
print(f"Monto total de gastos: ${monto_total:.2f}")


wb.save("informe_gastos.xlsx")
print("El informe de gastos se ha guardado en 'informe_gastos.xlsx'")